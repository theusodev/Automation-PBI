import tkinter as tk
from tkinter import filedialog
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
import configparser

# Carregar configurações anteriores (se existirem)
config = configparser.ConfigParser()
config_file = "config.ini"
if Path(config_file).exists():
    config.read(config_file)
    diretorio_origem_str = config.get("Configuracoes", "diretorio_origem_str")
    parte_titulo_arquivo = config.get("Configuracoes", "parte_titulo_arquivo")
    arquivo_destino_str = config.get("Configuracoes", "arquivo_destino_str")
    linha_inicial = config.getint("Configuracoes", "linha_inicial")
else:
    diretorio_origem_str = ""
    parte_titulo_arquivo = ""
    arquivo_destino_str = ""
    linha_inicial = 0

def salvar_configuracoes():
    config["Configuracoes"] = {
        "diretorio_origem_str": diretorio_origem_str,
        "parte_titulo_arquivo": parte_titulo_arquivo,
        "arquivo_destino_str": arquivo_destino_str,
        "linha_inicial": str(linha_inicial)
    }
    with open(config_file, "w") as f:
        config.write(f)

def executar_programa():
    global diretorio_origem_str, parte_titulo_arquivo, arquivo_destino_str, linha_inicial
    
    # Obter os valores dos inputs
    diretorio_origem_str = entrada_diretorio_origem.get()
    parte_titulo_arquivo = entrada_parte_titulo.get()
    arquivo_destino_str = entrada_arquivo_destino.get()

    try:
        linha_inicial = int(entrada_linha_inicial.get())
    except ValueError:
        resultado["text"] = "Por favor, insira um número válido para a linha inicial."
        return

    # Salvar as configurações
    salvar_configuracoes()

    diretorio_origem = Path(diretorio_origem_str)
    arquivo_destino = Path(arquivo_destino_str)

    if not diretorio_origem.exists():
        resultado["text"] = "O diretório origem não existe."
        return

    arquivo_mais_recente = None
    tempo_modificacao_mais_recente = float('-inf')

    for arquivo in diretorio_origem.glob('*'):
        if arquivo.is_file():
            if arquivo.suffix.lower() in ['.xlsx', '.xls', '.xlsm']:
                if parte_titulo_arquivo.lower() in arquivo.stem.lower():
                    tempo_modificacao = arquivo.stat().st_mtime
                    if tempo_modificacao > tempo_modificacao_mais_recente:
                        arquivo_mais_recente = arquivo
                        tempo_modificacao_mais_recente = tempo_modificacao

    if arquivo_mais_recente is not None:
        resultado["text"] = "Arquivo mais recente: " + str(arquivo_mais_recente)

        if tempo_modificacao_mais_recente > 0 and tempo_modificacao_mais_recente < 2**31:
            data_modificacao = datetime.fromtimestamp(tempo_modificacao_mais_recente)
            data_formatada = data_modificacao.strftime('%Y-%m-%d %H:%M:%S')
            resultado["text"] += "\nData modificação: " + data_formatada
        else:
            resultado["text"] += "\nTempo de modificação inválido."

        try:
            workbook = load_workbook(arquivo_mais_recente)
            planilha = workbook.worksheets[0]
            dados = []

            for linha in planilha.iter_rows(min_row=linha_inicial, values_only=True):
                if all(celula is None for celula in linha):
                    break
                dados.append(linha)

            dest_workbook = load_workbook(arquivo_destino)
            dest_planilha = dest_workbook.worksheets[0]

            ultima_linha_dest = dest_planilha.max_row
            while dest_planilha.cell(row=ultima_linha_dest, column=1).value is None and ultima_linha_dest > 1:
                ultima_linha_dest -= 1

            linha_destino = ultima_linha_dest + 1
            for linha in dados:
                for coluna, valor in enumerate(linha, start=1):
                    dest_planilha.cell(row=linha_destino, column=coluna, value=valor)

                linha_destino += 1

            dest_workbook.save(arquivo_destino)
            resultado["text"] += "\nDados copiados com sucesso para a planilha destino"

        except Exception as e:
            resultado["text"] += "\nOcorreu um erro ao copiar os dados para a planilha destino: " + str(e)

    else:
        resultado["text"] = f"Não foram encontrados arquivos que contenham '{parte_titulo_arquivo}' no diretório."

root = tk.Tk()
root.title("Programa de Processamento de Planilhas")

# Diretório de origem
tk.Label(root, text="Diretório de Origem:").grid(row=0, column=0)
entrada_diretorio_origem = tk.Entry(root)
entrada_diretorio_origem.grid(row=0, column=1)
entrada_diretorio_origem.insert(0, diretorio_origem_str)
botao_selecionar_diretorio = tk.Button(root, text="Selecionar Diretório", command=lambda: entrada_diretorio_origem.insert(tk.END, filedialog.askdirectory()))
botao_selecionar_diretorio.grid(row=0, column=2)

# Parte do título do arquivo
tk.Label(root, text="Parte do Título do Arquivo:").grid(row=1, column=0)
entrada_parte_titulo = tk.Entry(root)
entrada_parte_titulo.grid(row=1, column=1)
entrada_parte_titulo.insert(0, parte_titulo_arquivo)

# Arquivo de destino
tk.Label(root, text="Arquivo de Destino:").grid(row=2, column=0)
entrada_arquivo_destino = tk.Entry(root)
entrada_arquivo_destino.grid(row=2, column=1)
entrada_arquivo_destino.insert(0, arquivo_destino_str)
botao_selecionar_arquivo_destino = tk.Button(root, text="Selecionar Arquivo", command=lambda: entrada_arquivo_destino.insert(tk.END, filedialog.asksaveasfilename()))
botao_selecionar_arquivo_destino.grid(row=2, column=2)

# Linha inicial
tk.Label(root, text="Linha Inicial de Cópia:").grid(row=3, column=0)
entrada_linha_inicial = tk.Entry(root)
entrada_linha_inicial.grid(row=3, column=1)
entrada_linha_inicial.insert(0, linha_inicial)

# Botão para executar o programa
botao_executar = tk.Button(root, text="Executar Programa", command=executar_programa)
botao_executar.grid(row=4, columnspan=3)

# Resultado
resultado = tk.Label(root, text="")
resultado.grid(row=5, columnspan=3)

root.mainloop()
